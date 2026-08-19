#!/usr/bin/env python3
"""매출 데이터로 거래처별 청구서 시트·PDF를 생성한다."""

from __future__ import annotations

import argparse
import re
import shutil
import subprocess
import sys
import tempfile
from collections.abc import Callable
from datetime import date, datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

PROJECT_ROOT = (
    Path(sys.executable).resolve().parent
    if getattr(sys, "frozen", False)
    else Path(__file__).resolve().parent
)
GUI_INITIAL_DIR = r"C:\\"
DEFAULT_GUI_YEAR = "2025"

TEMPLATE_SHEET = "청구서"
PROTECTED_SHEETS = {"청구서", "청구서 BackUp", "청구서(sample)"}
SALES_SHEET = "매출"
CLIENT_SHEET = "거래처"

DETAIL_START_ROW = 20
TEMPLATE_DETAIL_ROWS = 16
SUMMARY_ROW_BASE = 36
DATE_NUMBER_FORMAT = "yy-mm-dd"
DATE_COLUMN_WIDTH = 12.0

SHEET_NAME_INVALID = re.compile(r"[\[\]\:\*\?/\\]")


class BillError(Exception):
    """청구서 생성 처리 중 발생하는 오류."""


LogFn = Callable[[str], None] | None


def emit_log(log_fn: LogFn, message: str, *, error: bool = False) -> None:
    if log_fn is not None:
        log_fn(message)
    elif error:
        print(message, file=sys.stderr)
    else:
        print(message)


def resolve_path(path_str: str) -> Path:
    path = Path(path_str)
    if not path.is_absolute():
        path = PROJECT_ROOT / path
    return path.resolve()


def parse_year_month(value: str) -> tuple[int, int]:
    match = re.fullmatch(r"(\d{4})-(\d{2})", value.strip())
    if not match:
        raise BillError(f"연월 형식이 올바르지 않습니다 (yyyy-mm 필요): {value}")
    year = int(match.group(1))
    month = int(match.group(2))
    if not 1 <= month <= 12:
        raise BillError(f"월은 01~12 범위여야 합니다: {value}")
    return year, month


def to_yymm(year: int, month: int) -> str:
    return f"{year % 100:02d}{month:02d}"


def normalize_code(value: object) -> int:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        raise BillError("거래처코드가 비어 있습니다.")
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        number = float(value)
        if number.is_integer():
            return int(number)
        raise BillError(f"거래처코드 형식 오류: {value}")
    text = str(value).strip()
    if not text:
        raise BillError("거래처코드가 비어 있습니다.")
    number = float(text)
    if number.is_integer():
        return int(number)
    raise BillError(f"거래처코드 형식 오류: {value}")


def sanitize_sheet_name(name: str) -> str:
    cleaned = SHEET_NAME_INVALID.sub("", name)
    return cleaned[:31]


def build_sheet_name(client_name: str, client_code: int, yymm: str) -> str:
    return sanitize_sheet_name(f"{client_name}_{client_code}_{yymm}")


def to_date(value: object) -> date | datetime:
    if isinstance(value, datetime):
        return value.date() if hasattr(value, "date") else value
    if isinstance(value, date):
        return value
    if pd.isna(value):
        raise BillError("매출일이 비어 있습니다.")
    parsed = pd.to_datetime(value)
    return parsed.date()


def to_number(value: object) -> int | float:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        raise BillError("숫자 값이 비어 있습니다.")
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if isinstance(value, float) and value.is_integer():
            return int(value)
        return value
    text = str(value).strip()
    number = float(text)
    if number.is_integer():
        return int(number)
    return number


def load_clients(sales_xlsx: Path) -> dict[int, dict[str, object]]:
    df = pd.read_excel(sales_xlsx, sheet_name=CLIENT_SHEET)
    clients: dict[int, dict[str, object]] = {}
    for _, row in df.iterrows():
        code = normalize_code(row["거래처코드"])
        clients[code] = {
            "name": str(row["거래처명"]),
            "zip": row["우편번호"],
            "address": str(row["주소"]),
        }
    return clients


def load_sales_groups(
    sales_xlsx: Path, year: int, month: int
) -> dict[int, list[dict[str, object]]]:
    df = pd.read_excel(sales_xlsx, sheet_name=SALES_SHEET)
    if "매출일" not in df.columns:
        raise BillError(f"'{SALES_SHEET}' 시트에 '매출일' 열이 없습니다.")

    df["매출일"] = pd.to_datetime(df["매출일"])
    filtered = df[
        (df["매출일"].dt.year == year) & (df["매출일"].dt.month == month)
    ]

    groups: dict[int, list[dict[str, object]]] = {}
    for _, row in filtered.iterrows():
        code = normalize_code(row["거래처코드"])
        groups.setdefault(code, []).append(
            {
                "매출일": row["매출일"],
                "상품명": row["상품명"],
                "매출수량": row["매출수량"],
                "단가": row["단가"],
            }
        )
    return groups


def _merged_ranges(ws) -> set[str]:
    return {str(item) for item in ws.merged_cells.ranges}


def ensure_detail_row_merge(ws, row: int) -> None:
    merge_range = f"B{row}:E{row}"
    if merge_range not in _merged_ranges(ws):
        ws.merge_cells(merge_range)


def clear_detail_row(ws, row: int) -> None:
    for col in (1, 2, 6, 7, 8, 9):
        ws.cell(row=row, column=col).value = None


def fill_invoice_sheet(ws, client: dict[str, object], details: list[dict[str, object]]) -> None:
    ws["B5"] = client["zip"]
    ws["B6"] = client["address"]
    ws["B7"] = client["name"]

    detail_count = len(details)
    if detail_count == 0:
        raise BillError("明細 데이터가 없습니다.")

    if detail_count > TEMPLATE_DETAIL_ROWS:
        ws.insert_rows(SUMMARY_ROW_BASE, detail_count - TEMPLATE_DETAIL_ROWS)

    last_detail_row = DETAIL_START_ROW + detail_count - 1
    summary_row = SUMMARY_ROW_BASE + max(0, detail_count - TEMPLATE_DETAIL_ROWS)
    tax_row = summary_row + 1

    for index, detail in enumerate(details):
        row = DETAIL_START_ROW + index
        ensure_detail_row_merge(ws, row)
        date_cell = ws.cell(row=row, column=1)
        date_cell.value = to_date(detail["매출일"])
        date_cell.number_format = DATE_NUMBER_FORMAT
        ws.cell(row=row, column=2).value = str(detail["상품명"])
        ws.cell(row=row, column=6).value = to_number(detail["매출수량"])
        ws.cell(row=row, column=7).value = to_number(detail["단가"])
        ws.cell(row=row, column=8).value = f"=F{row}*G{row}"
        ws.cell(row=row, column=9).value = None

    current_width = ws.column_dimensions["A"].width or 0
    if current_width < DATE_COLUMN_WIDTH:
        ws.column_dimensions["A"].width = DATE_COLUMN_WIDTH

    if detail_count < TEMPLATE_DETAIL_ROWS:
        for row in range(DETAIL_START_ROW + detail_count, SUMMARY_ROW_BASE):
            clear_detail_row(ws, row)

    ws.cell(row=summary_row, column=8).value = (
        f"=SUM(H{DETAIL_START_ROW}:H{last_detail_row})"
    )
    ws.cell(row=tax_row, column=8).value = f"=INT(H{summary_row}*0.1)"
    ws["E17"] = f"=H{summary_row}"
    ws["G17"] = f"=H{tax_row}"
    ws["I17"] = f"=H{summary_row}+H{tax_row}"
    ws.print_area = f"A1:I{tax_row}"


def create_invoice_sheets(
    invoice_xlsx: Path,
    clients: dict[int, dict[str, object]],
    sales_groups: dict[int, list[dict[str, object]]],
    yymm: str,
    *,
    log_fn: LogFn = None,
) -> tuple[list[str], list[str]]:
    try:
        workbook = load_workbook(invoice_xlsx)
    except PermissionError as exc:
        raise BillError(
            f"청구서 파일을 열 수 없습니다. 파일이 열려 있는지 확인하세요: {invoice_xlsx}"
        ) from exc

    if TEMPLATE_SHEET not in workbook.sheetnames:
        workbook.close()
        raise BillError(f"시트를 찾을 수 없습니다: {TEMPLATE_SHEET}")

    template = workbook[TEMPLATE_SHEET]
    created: list[str] = []
    skipped: list[str] = []
    total = len(sales_groups)

    emit_log(log_fn, f"청구서 시트 생성 시작 (대상 거래처 {total}건)")

    for index, (client_code, details) in enumerate(sorted(sales_groups.items()), start=1):
        client = clients.get(client_code)
        if client is None:
            skipped.append(str(client_code))
            emit_log(
                log_fn,
                f"경고: 거래처코드 {client_code} — 거래처 시트에 없어 skip합니다.",
                error=True,
            )
            continue

        sheet_name = build_sheet_name(str(client["name"]), client_code, yymm)
        if sheet_name in workbook.sheetnames:
            del workbook[sheet_name]

        new_sheet = workbook.copy_worksheet(template)
        new_sheet.title = sheet_name
        fill_invoice_sheet(new_sheet, client, details)
        created.append(sheet_name)
        emit_log(
            log_fn,
            f"시트 생성 ({index}/{total}): {sheet_name} (明細 {len(details)}행)",
        )

    emit_log(log_fn, "청구서.xlsx 저장 중...")
    try:
        workbook.save(invoice_xlsx)
    except PermissionError as exc:
        workbook.close()
        raise BillError(
            f"청구서 파일을 저장할 수 없습니다. 파일을 닫은 뒤 다시 실행하세요: {invoice_xlsx}"
        ) from exc
    finally:
        workbook.close()

    emit_log(log_fn, f"청구서 시트 생성 완료: {len(created)}건")
    return created, skipped


def find_libreoffice() -> Path | None:
    candidates = [
        Path(r"C:\Program Files\LibreOffice\program\soffice.exe"),
        Path(r"C:\Program Files (x86)\LibreOffice\program\soffice.exe"),
    ]
    for candidate in candidates:
        if candidate.is_file():
            return candidate
    return None


def _verify_pdf_file(pdf_path: Path) -> None:
    if not pdf_path.is_file():
        raise BillError(f"PDF 파일이 생성되지 않았습니다: {pdf_path}")
    if pdf_path.stat().st_size <= 0:
        raise BillError(f"PDF 파일이 비어 있습니다: {pdf_path}")


def _create_single_sheet_workbook(
    invoice_xlsx: Path, sheet_name: str, target_xlsx: Path
) -> None:
    workbook = load_workbook(invoice_xlsx)
    if sheet_name not in workbook.sheetnames:
        workbook.close()
        raise BillError(f"PDF 변환 대상 시트를 찾을 수 없습니다: {sheet_name}")

    for name in list(workbook.sheetnames):
        if name != sheet_name:
            del workbook[name]

    target_xlsx.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(target_xlsx)
    workbook.close()


def export_pdfs_with_excel_com(
    invoice_xlsx: Path,
    pdf_dir: Path,
    sheet_names: list[str],
    *,
    log_fn: LogFn = None,
) -> None:
    import pythoncom
    import win32com.client

    pdf_dir.mkdir(parents=True, exist_ok=True)
    invoice_path = str(invoice_xlsx.resolve())
    pythoncom.CoInitialize()
    excel = None

    try:
        if getattr(sys, "frozen", False):
            excel = win32com.client.Dispatch("Excel.Application")
        else:
            try:
                excel = win32com.client.gencache.EnsureDispatch("Excel.Application")
            except Exception:
                excel = win32com.client.Dispatch("Excel.Application")

        excel.Visible = False
        excel.DisplayAlerts = False
        excel.ScreenUpdating = False

        workbook = excel.Workbooks.Open(invoice_path, ReadOnly=True)
        try:
            total = len(sheet_names)
            for index, sheet_name in enumerate(sheet_names, start=1):
                worksheet = workbook.Worksheets(sheet_name)
                pdf_path = (pdf_dir / f"{sheet_name}.pdf").resolve()
                emit_log(log_fn, f"PDF 출력 ({index}/{total}): {pdf_path.name}")
                worksheet.ExportAsFixedFormat(
                    Type=0,
                    Filename=str(pdf_path),
                    Quality=0,
                    IncludeDocProperties=True,
                    IgnorePrintAreas=False,
                    OpenAfterPublish=False,
                )
                _verify_pdf_file(pdf_path)
        finally:
            workbook.Close(SaveChanges=False)
    finally:
        if excel is not None:
            excel.Quit()
        pythoncom.CoUninitialize()


def export_pdfs_with_libreoffice(
    invoice_xlsx: Path,
    pdf_dir: Path,
    sheet_names: list[str],
    *,
    log_fn: LogFn = None,
) -> None:
    soffice = find_libreoffice()
    if soffice is None:
        raise BillError("LibreOffice(soffice.exe)를 찾을 수 없습니다.")

    pdf_dir.mkdir(parents=True, exist_ok=True)
    emit_log(log_fn, f"LibreOffice PDF 변환 시작 → {pdf_dir}")

    with tempfile.TemporaryDirectory(prefix="bill_pdf_") as temp_dir_name:
        temp_dir = Path(temp_dir_name)
        total = len(sheet_names)
        for index, sheet_name in enumerate(sheet_names, start=1):
            temp_xlsx = temp_dir / f"{sheet_name}.xlsx"
            _create_single_sheet_workbook(invoice_xlsx, sheet_name, temp_xlsx)
            emit_log(log_fn, f"LibreOffice 변환 ({index}/{total}): {sheet_name}.pdf")

            result = subprocess.run(
                [
                    str(soffice),
                    "--headless",
                    "--nologo",
                    "--convert-to",
                    "pdf",
                    "--outdir",
                    str(temp_dir),
                    str(temp_xlsx),
                ],
                capture_output=True,
                text=True,
                timeout=180,
                check=False,
            )
            if result.returncode != 0:
                detail = (result.stderr or result.stdout or "").strip()
                raise BillError(
                    f"LibreOffice PDF 변환 실패: {sheet_name}"
                    + (f" ({detail})" if detail else "")
                )

            generated_pdf = temp_dir / f"{temp_xlsx.stem}.pdf"
            target_pdf = pdf_dir / f"{sheet_name}.pdf"
            if not generated_pdf.is_file():
                raise BillError(f"LibreOffice PDF 파일이 생성되지 않았습니다: {sheet_name}")

            if target_pdf.exists():
                target_pdf.unlink()
            shutil.move(str(generated_pdf), str(target_pdf))
            _verify_pdf_file(target_pdf)


def _register_reportlab_korean_font() -> str:
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    font_name = "MalgunGothic"
    if font_name in pdfmetrics.getRegisteredFontNames():
        return font_name

    font_path = Path(r"C:\Windows\Fonts\malgun.ttf")
    if not font_path.is_file():
        raise BillError(
            "한글 PDF 폰트(malgun.ttf)를 찾을 수 없습니다. Windows 기본 폰트를 확인하세요."
        )

    pdfmetrics.registerFont(TTFont(font_name, str(font_path)))
    return font_name


def _read_invoice_details_from_sheet(ws) -> tuple[list[dict[str, object]], int, int, int]:
    details: list[dict[str, object]] = []
    summary_labels = {"과세대상액", "과세대상소비세액"}

    for row in range(DETAIL_START_ROW, ws.max_row + 1):
        product = ws.cell(row=row, column=2).value
        if product in summary_labels:
            break
        if product is None or str(product).strip() == "":
            continue

        sale_date = ws.cell(row=row, column=1).value
        qty = ws.cell(row=row, column=6).value
        unit_price = ws.cell(row=row, column=7).value
        if qty is None or unit_price is None:
            continue

        amount = float(qty) * float(unit_price)
        details.append(
            {
                "date": sale_date,
                "product": str(product),
                "qty": qty,
                "unit_price": unit_price,
                "amount": amount,
            }
        )

    subtotal = int(round(sum(float(item["amount"]) for item in details)))
    tax = int(subtotal * 0.1)
    total = subtotal + tax
    return details, subtotal, tax, total


def export_pdfs_with_reportlab(
    invoice_xlsx: Path,
    pdf_dir: Path,
    sheet_names: list[str],
    *,
    log_fn: LogFn = None,
) -> None:
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas

    font_name = _register_reportlab_korean_font()
    pdf_dir.mkdir(parents=True, exist_ok=True)
    emit_log(log_fn, f"내장 PDF 변환 시작 → {pdf_dir}")

    workbook = load_workbook(invoice_xlsx, data_only=False)
    try:
        total = len(sheet_names)
        for index, sheet_name in enumerate(sheet_names, start=1):
            if sheet_name not in workbook.sheetnames:
                raise BillError(f"PDF 변환 대상 시트를 찾을 수 없습니다: {sheet_name}")

            ws = workbook[sheet_name]
            details, subtotal, tax, total_amount = _read_invoice_details_from_sheet(ws)
            pdf_path = (pdf_dir / f"{sheet_name}.pdf").resolve()
            emit_log(log_fn, f"PDF 생성 ({index}/{total}): {pdf_path.name}")

            pdf = canvas.Canvas(str(pdf_path), pagesize=A4)
            width, height = A4
            y = height - 40

            pdf.setFont(font_name, 16)
            pdf.drawCentredString(width / 2, y, "청구서")
            y -= 30

            pdf.setFont(font_name, 10)
            pdf.drawString(40, y, f"우편번호: {ws['B5'].value or ''}")
            y -= 16
            pdf.drawString(40, y, f"주소: {ws['B6'].value or ''}")
            y -= 16
            pdf.drawString(40, y, f"거래처: {ws['B7'].value or ''}")
            y -= 24

            pdf.setFont(font_name, 9)
            pdf.drawString(40, y, "날짜")
            pdf.drawString(100, y, "상품명")
            pdf.drawString(320, y, "수량")
            pdf.drawString(380, y, "단가")
            pdf.drawString(450, y, "금액")
            y -= 12
            pdf.line(40, y, width - 40, y)
            y -= 14

            for detail in details:
                if y < 120:
                    pdf.showPage()
                    pdf.setFont(font_name, 9)
                    y = height - 40

                date_text = detail["date"]
                if hasattr(date_text, "strftime"):
                    date_text = date_text.strftime("%y-%m-%d")
                else:
                    date_text = str(date_text or "")

                pdf.drawString(40, y, date_text)
                pdf.drawString(100, y, str(detail["product"])[:28])
                pdf.drawRightString(360, y, f"{detail['qty']}")
                pdf.drawRightString(430, y, f"{int(float(detail['unit_price'])):,}")
                pdf.drawRightString(width - 40, y, f"{int(float(detail['amount'])):,}")
                y -= 14

            y -= 10
            pdf.line(40, y, width - 40, y)
            y -= 18
            pdf.drawString(320, y, "과세대상액")
            pdf.drawRightString(width - 40, y, f"{subtotal:,}")
            y -= 16
            pdf.drawString(320, y, "소비세")
            pdf.drawRightString(width - 40, y, f"{tax:,}")
            y -= 16
            pdf.setFont(font_name, 10)
            pdf.drawString(320, y, "청구액")
            pdf.drawRightString(width - 40, y, f"{total_amount:,}")

            pdf.save()
            _verify_pdf_file(pdf_path)
    finally:
        workbook.close()


def export_pdfs(
    invoice_xlsx: Path,
    pdf_dir: Path,
    sheet_names: list[str],
    *,
    log_fn: LogFn = None,
) -> None:
    if not sheet_names:
        return

    pdf_dir.mkdir(parents=True, exist_ok=True)
    emit_log(log_fn, f"PDF 출력 시작 → {pdf_dir}")

    errors: list[str] = []
    try:
        export_pdfs_with_excel_com(
            invoice_xlsx, pdf_dir, sheet_names, log_fn=log_fn
        )
        emit_log(log_fn, f"PDF 출력 완료 (Excel COM): {len(sheet_names)}개")
        return
    except Exception as exc:
        errors.append(f"Excel COM: {exc}")
        emit_log(
            log_fn,
            f"Excel COM PDF 출력 실패, LibreOffice 방식으로 재시도합니다: {exc}",
            error=True,
        )

    try:
        export_pdfs_with_libreoffice(
            invoice_xlsx, pdf_dir, sheet_names, log_fn=log_fn
        )
        emit_log(log_fn, f"PDF 출력 완료 (LibreOffice): {len(sheet_names)}개")
        return
    except Exception as exc:
        errors.append(f"LibreOffice: {exc}")
        emit_log(log_fn, f"LibreOffice PDF 출력 실패: {exc}", error=True)

    try:
        export_pdfs_with_reportlab(
            invoice_xlsx, pdf_dir, sheet_names, log_fn=log_fn
        )
        emit_log(log_fn, f"PDF 출력 완료 (내장 변환): {len(sheet_names)}개")
        return
    except Exception as exc:
        errors.append(f"내장 PDF 변환: {exc}")
        emit_log(log_fn, f"내장 PDF 변환 실패: {exc}", error=True)

    raise BillError(
        "PDF 출력에 실패했습니다. Excel/LibreOffice 사용 또는 내장 PDF 변환을 확인하세요.\n"
        + "\n".join(errors)
    )


def delete_created_sheets(
    invoice_xlsx: Path,
    sheet_names: list[str],
    *,
    log_fn: LogFn = None,
) -> None:
    if not sheet_names:
        return

    emit_log(log_fn, f"임시 시트 삭제 시작: {len(sheet_names)}건")
    workbook = load_workbook(invoice_xlsx)
    deleted = 0
    for sheet_name in sheet_names:
        if sheet_name in workbook.sheetnames and sheet_name not in PROTECTED_SHEETS:
            del workbook[sheet_name]
            deleted += 1

    try:
        workbook.save(invoice_xlsx)
    except PermissionError as exc:
        workbook.close()
        raise BillError(
            f"청구서 파일을 저장할 수 없습니다. 파일을 닫은 뒤 다시 실행하세요: {invoice_xlsx}"
        ) from exc
    finally:
        workbook.close()

    if deleted != len(sheet_names):
        raise BillError("생성 시트 삭제 중 일부 시트를 찾지 못했습니다.")

    emit_log(log_fn, f"임시 시트 삭제 완료: {deleted}건")


def format_summary(
    year_month: str,
    client_count: int,
    detail_count: int,
    pdf_count: int,
    deleted_count: int,
    pdf_dir: Path,
    invoice_xlsx: Path,
    *,
    dry_run: bool,
    skipped_count: int,
) -> str:
    lines = [
        f"대상 월: {year_month}",
        f"거래처: {client_count}건",
        f"明細 행: {detail_count}건",
    ]
    if skipped_count:
        lines.append(f"skip: {skipped_count}건 (거래처 미매칭)")
    if dry_run:
        lines.append("dry-run: 시트 생성·PDF·삭제를 수행하지 않았습니다.")
        return "\n".join(lines)
    lines.extend(
        [
            f"PDF 출력: {pdf_count}개 → {pdf_dir}",
            f"삭제 시트: {deleted_count}개",
            f"저장 완료: {invoice_xlsx}",
        ]
    )
    return "\n".join(lines)


def print_summary(
    year_month: str,
    client_count: int,
    detail_count: int,
    pdf_count: int,
    deleted_count: int,
    pdf_dir: Path,
    invoice_xlsx: Path,
    *,
    dry_run: bool,
    skipped_count: int,
    log_fn: LogFn = None,
) -> str:
    summary = format_summary(
        year_month,
        client_count,
        detail_count,
        pdf_count,
        deleted_count,
        pdf_dir,
        invoice_xlsx,
        dry_run=dry_run,
        skipped_count=skipped_count,
    )
    if log_fn is not None:
        for line in summary.splitlines():
            log_fn(line)
    else:
        print(summary)
    return summary


def run(
    year_month: str,
    sales_xlsx: Path,
    invoice_xlsx: Path,
    pdf_dir: Path,
    *,
    dry_run: bool = False,
    log_fn: LogFn = None,
) -> str:
    if not sales_xlsx.is_file():
        raise BillError(f"매출 파일이 존재하지 않습니다: {sales_xlsx}")
    if not invoice_xlsx.is_file():
        raise BillError(f"청구서 파일이 존재하지 않습니다: {invoice_xlsx}")

    year, month = parse_year_month(year_month)
    yymm = to_yymm(year, month)

    emit_log(log_fn, f"작업 시작: 대상 월 {year_month}")
    emit_log(log_fn, "매출·거래처 데이터 로드 중...")
    clients = load_clients(sales_xlsx)
    sales_groups = load_sales_groups(sales_xlsx, year, month)
    emit_log(log_fn, f"매출 데이터 로드 완료: 거래처 {len(sales_groups)}건")

    if not sales_groups:
        emit_log(log_fn, "해당 월 매출 데이터가 없습니다.")
        return print_summary(
            year_month,
            0,
            0,
            0,
            0,
            pdf_dir,
            invoice_xlsx,
            dry_run=dry_run,
            skipped_count=0,
            log_fn=log_fn,
        )

    matched_groups = {
        code: sales_groups[code] for code in sales_groups if code in clients
    }
    skipped_count = len(sales_groups) - len(matched_groups)
    detail_count = sum(len(items) for items in matched_groups.values())

    if dry_run:
        emit_log(log_fn, "dry-run: 시트 생성·PDF·삭제를 수행하지 않습니다.")
        return print_summary(
            year_month,
            len(matched_groups),
            detail_count,
            0,
            0,
            pdf_dir,
            invoice_xlsx,
            dry_run=True,
            skipped_count=skipped_count,
            log_fn=log_fn,
        )

    created_sheets, _ = create_invoice_sheets(
        invoice_xlsx, clients, sales_groups, yymm, log_fn=log_fn
    )
    export_pdfs(invoice_xlsx, pdf_dir, created_sheets, log_fn=log_fn)
    delete_created_sheets(invoice_xlsx, created_sheets, log_fn=log_fn)

    emit_log(log_fn, "모든 작업이 완료되었습니다.")
    return print_summary(
        year_month,
        len(created_sheets),
        detail_count,
        len(created_sheets),
        len(created_sheets),
        pdf_dir,
        invoice_xlsx,
        dry_run=False,
        skipped_count=skipped_count,
        log_fn=log_fn,
    )


def launch_bill_gui(*, dry_run: bool = False) -> int:
    """매출·청구서·PDF 경로와 연·월을 선택하는 GUI."""
    import tkinter as tk
    from tkinter import filedialog, messagebox, scrolledtext, ttk

    root = tk.Tk()
    root.title("거래처별 청구서 PDF 생성")
    root.resizable(True, True)
    root.minsize(640, 520)

    sales_path_var = tk.StringVar()
    invoice_path_var = tk.StringVar()
    pdf_dir_var = tk.StringVar()
    year_var = tk.StringVar(value=DEFAULT_GUI_YEAR)
    month_var = tk.StringVar(value=f"{datetime.now().month:02d}")

    frame = ttk.Frame(root, padding=12)
    frame.grid(row=0, column=0, sticky="nsew")
    root.columnconfigure(0, weight=1)
    root.rowconfigure(0, weight=1)
    frame.rowconfigure(6, weight=1)
    frame.columnconfigure(1, weight=1)

    def append_log(message: str) -> None:
        log_text.configure(state="normal")
        log_text.insert("end", message + "\n")
        log_text.configure(state="disabled")
        log_text.see("end")
        root.update_idletasks()

    def browse_sales() -> None:
        path = filedialog.askopenfilename(
            title="매출.xlsx 파일 선택",
            initialdir=GUI_INITIAL_DIR,
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        )
        if path:
            sales_path_var.set(path)

    def browse_invoice() -> None:
        path = filedialog.askopenfilename(
            title="청구서.xlsx 파일 선택",
            initialdir=GUI_INITIAL_DIR,
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        )
        if path:
            invoice_path_var.set(path)

    def browse_pdf_dir() -> None:
        path = filedialog.askdirectory(
            title="PDF 저장 폴더 선택",
            initialdir=GUI_INITIAL_DIR,
        )
        if path:
            pdf_dir_var.set(path)

    def validate_inputs() -> tuple[str, Path, Path, Path] | None:
        sales_text = sales_path_var.get().strip()
        invoice_text = invoice_path_var.get().strip()
        pdf_text = pdf_dir_var.get().strip()
        year_text = year_var.get().strip()
        month_text = month_var.get().strip()

        if not sales_text:
            messagebox.showerror("입력 오류", "매출.xlsx 파일을 선택하세요.", parent=root)
            return None
        if not invoice_text:
            messagebox.showerror("입력 오류", "청구서.xlsx 파일을 선택하세요.", parent=root)
            return None
        if not pdf_text:
            messagebox.showerror("입력 오류", "PDF 저장 폴더를 선택하세요.", parent=root)
            return None

        try:
            year = int(year_text)
            month = int(month_text)
            if not 1 <= month <= 12:
                raise ValueError
            year_month = f"{year}-{month:02d}"
        except ValueError:
            messagebox.showerror(
                "입력 오류",
                "연도는 숫자, 월은 01~12 범위로 입력하세요.",
                parent=root,
            )
            return None

        sales_path = Path(sales_text).resolve()
        invoice_path = Path(invoice_text).resolve()
        pdf_dir = Path(pdf_text).resolve()

        if not sales_path.is_file():
            messagebox.showerror(
                "입력 오류",
                f"매출 파일이 존재하지 않습니다:\n{sales_path}",
                parent=root,
            )
            return None
        if not invoice_path.is_file():
            messagebox.showerror(
                "입력 오류",
                f"청구서 파일이 존재하지 않습니다:\n{invoice_path}",
                parent=root,
            )
            return None

        return year_month, sales_path, invoice_path, pdf_dir

    def on_run() -> None:
        validated = validate_inputs()
        if validated is None:
            return

        year_month, sales_path, invoice_path, pdf_dir = validated
        run_button.config(state="disabled")
        log_text.configure(state="normal")
        log_text.delete("1.0", "end")
        log_text.configure(state="disabled")
        append_log("=" * 50)
        append_log("작업을 시작합니다.")
        root.update_idletasks()

        try:
            run(
                year_month,
                sales_path,
                invoice_path,
                pdf_dir,
                dry_run=dry_run,
                log_fn=append_log,
            )
        except BillError as exc:
            append_log(f"오류: {exc}")
        finally:
            run_button.config(state="normal")

    ttk.Label(frame, text="매출.xlsx").grid(row=0, column=0, sticky="e", padx=(0, 8), pady=4)
    ttk.Entry(frame, textvariable=sales_path_var, width=52).grid(
        row=0, column=1, sticky="we", pady=4
    )
    ttk.Button(frame, text="찾아보기", command=browse_sales, width=10).grid(
        row=0, column=2, padx=(8, 0), pady=4
    )

    ttk.Label(frame, text="청구서.xlsx").grid(row=1, column=0, sticky="e", padx=(0, 8), pady=4)
    ttk.Entry(frame, textvariable=invoice_path_var, width=52).grid(
        row=1, column=1, sticky="we", pady=4
    )
    ttk.Button(frame, text="찾아보기", command=browse_invoice, width=10).grid(
        row=1, column=2, padx=(8, 0), pady=4
    )

    ttk.Label(frame, text="매출 연도").grid(row=2, column=0, sticky="e", padx=(0, 8), pady=4)
    ttk.Entry(frame, textvariable=year_var, width=10).grid(row=2, column=1, sticky="w", pady=4)

    ttk.Label(frame, text="매출 월").grid(row=3, column=0, sticky="e", padx=(0, 8), pady=4)
    month_combo = ttk.Combobox(
        frame,
        textvariable=month_var,
        values=[f"{m:02d}" for m in range(1, 13)],
        width=8,
        state="readonly",
    )
    month_combo.grid(row=3, column=1, sticky="w", pady=4)

    ttk.Label(frame, text="PDF 저장 폴더").grid(
        row=4, column=0, sticky="e", padx=(0, 8), pady=4
    )
    ttk.Entry(frame, textvariable=pdf_dir_var, width=52).grid(
        row=4, column=1, sticky="we", pady=4
    )
    ttk.Button(frame, text="찾아보기", command=browse_pdf_dir, width=10).grid(
        row=4, column=2, padx=(8, 0), pady=4
    )

    ttk.Label(frame, text="작업 로그").grid(
        row=5, column=0, sticky="nw", padx=(0, 8), pady=(8, 4)
    )
    log_text = scrolledtext.ScrolledText(
        frame,
        width=70,
        height=12,
        state="disabled",
        wrap="word",
    )
    log_text.grid(row=6, column=0, columnspan=3, sticky="nsew", pady=(0, 8))

    button_frame = ttk.Frame(frame)
    button_frame.grid(row=7, column=0, columnspan=3, pady=(4, 0))
    run_button = ttk.Button(
        button_frame,
        text="실행" if not dry_run else "검증(dry-run)",
        command=on_run,
        width=14,
    )
    run_button.grid(row=0, column=0, padx=4)
    ttk.Button(button_frame, text="닫기", command=root.destroy, width=10).grid(
        row=0, column=1, padx=4
    )

    root.mainloop()
    return 0


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="매출 데이터로 거래처별 청구서 시트·PDF를 생성합니다."
    )
    parser.add_argument(
        "-m",
        "--year-month",
        default=None,
        help="대상 연월 (yyyy-mm, 예: 2025-08). --no-gui 사용 시 필수",
    )
    parser.add_argument(
        "--no-gui",
        action="store_true",
        help="연·월 선택 대화상자 없이 CLI 인자만 사용",
    )
    parser.add_argument(
        "--sales-xlsx",
        default="resData/매출.xlsx",
        help="매출 Excel 경로 (기본값: resData/매출.xlsx)",
    )
    parser.add_argument(
        "--invoice-xlsx",
        default="resData/청구서.xlsx",
        help="청구서 Excel 경로 (기본값: resData/청구서.xlsx)",
    )
    parser.add_argument(
        "--pdf-dir",
        default="resData/PDF",
        help="PDF 출력 폴더 (기본값: resData/PDF)",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="시트/PDF 생성 없이 대상·건수만 출력",
    )
    return parser


def main(argv: list[str] | None = None) -> int:
    parser = build_parser()
    args = parser.parse_args(argv)

    if args.no_gui:
        if not args.year_month:
            parser.error("--no-gui 사용 시 -m/--year-month 가 필요합니다.")
        try:
            run(
                args.year_month,
                resolve_path(args.sales_xlsx),
                resolve_path(args.invoice_xlsx),
                resolve_path(args.pdf_dir),
                dry_run=args.dry_run,
            )
        except BillError as exc:
            print(f"오류: {exc}", file=sys.stderr)
            return 1
        return 0

    return launch_bill_gui(dry_run=args.dry_run)


if __name__ == "__main__":
    raise SystemExit(main())
