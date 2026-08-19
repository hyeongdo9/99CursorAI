#!/usr/bin/env python3
"""CSV 파일을 매출.xlsx의 매출 시트에 적재한다."""

from __future__ import annotations

import argparse
import shutil
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

PROJECT_ROOT = (
    Path(sys.executable).resolve().parent
    if getattr(sys, "frozen", False)
    else Path(__file__).resolve().parent
)
GUI_INITIAL_DIR = r"C:\\"

EXPECTED_COLUMNS = [
    "매출일",
    "영업소",
    "영업담당",
    "기간",
    "전표번호",
    "상품코드",
    "상품명",
    "대분류",
    "중분류",
    "소분류",
    "거래처코드",
    "거래처명",
    "매출수량",
    "매출금액",
    "매출이익",
    "단가",
]

NUMERIC_COLUMNS = {
    "기간",
    "전표번호",
    "상품코드",
    "거래처코드",
    "매출수량",
    "매출금액",
    "매출이익",
    "단가",
}

AMOUNT_COLUMNS = {"매출금액", "매출이익", "단가"}
AMOUNT_NUMBER_FORMAT = "#,##0"


class CsvImportError(Exception):
    """CSV 적재 처리 중 발생하는 오류."""


def select_paths_via_gui() -> tuple[Path | None, Path | None]:
    """tkinter 대화 상자로 CSV 폴더와 매출.xlsx 경로를 선택한다."""
    import tkinter as tk
    from tkinter import filedialog

    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)

    csv_dir_str = filedialog.askdirectory(
        title="CSV 원본 폴더 선택",
        initialdir=GUI_INITIAL_DIR,
    )
    if not csv_dir_str:
        root.destroy()
        return None, None

    xlsx_path_str = filedialog.askopenfilename(
        title="매출.xlsx 파일 선택",
        initialdir=GUI_INITIAL_DIR,
        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
    )
    root.destroy()

    if not xlsx_path_str:
        return None, None

    return Path(csv_dir_str).resolve(), Path(xlsx_path_str).resolve()


def resolve_path(path_str: str) -> Path:
    path = Path(path_str)
    if not path.is_absolute():
        path = PROJECT_ROOT / path
    return path.resolve()


def collect_csv_files(csv_dir: Path) -> tuple[list[Path], list[Path]]:
    if not csv_dir.is_dir():
        raise CsvImportError(f"CSV 폴더가 존재하지 않습니다: {csv_dir}")

    pending = sorted(csv_dir.glob("*.csv"), key=lambda p: p.name)
    completed_dir = csv_dir / "completed"
    completed: list[Path] = []
    if completed_dir.is_dir():
        completed = sorted(completed_dir.glob("*.csv"), key=lambda p: p.name)

    return pending, completed


def read_and_merge_csv_files(csv_files: list[Path]) -> pd.DataFrame:
    if not csv_files:
        raise CsvImportError("처리할 CSV 파일이 없습니다.")

    frames: list[pd.DataFrame] = []
    expected_header: list[str] | None = None

    for csv_path in csv_files:
        try:
            df = pd.read_csv(csv_path, encoding="utf-8-sig")
        except OSError as exc:
            raise CsvImportError(f"CSV 파일을 읽을 수 없습니다: {csv_path.name}") from exc

        header = [str(col) for col in df.columns.tolist()]
        if expected_header is None:
            if header != EXPECTED_COLUMNS:
                raise CsvImportError(
                    f"헤더가 기대 형식과 다릅니다: {csv_path.name}\n"
                    f"  기대: {EXPECTED_COLUMNS}\n"
                    f"  실제: {header}"
                )
            expected_header = EXPECTED_COLUMNS
        elif header != expected_header:
            raise CsvImportError(
                f"헤더가 기준 CSV와 다릅니다: {csv_path.name}\n"
                f"  기준: {expected_header}\n"
                f"  실제: {header}"
            )

        frames.append(df)

    return pd.concat(frames, ignore_index=True)


def _to_number(value: object) -> int | float | None:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if isinstance(value, float) and value.is_integer():
            return int(value)
        return value
    text = str(value).strip()
    if not text:
        return None
    number = float(text)
    if number.is_integer():
        return int(number)
    return number


def _to_date(value: object):
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    if hasattr(value, "year") and hasattr(value, "month"):
        return value
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%Y/%m/%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    raise CsvImportError(f"날짜 형식을 해석할 수 없습니다: {text}")


def _cell_value(column_name: str, value: object):
    if pd.isna(value):
        return None
    if column_name == "매출일":
        return _to_date(value)
    if column_name in NUMERIC_COLUMNS:
        return _to_number(value)
    return str(value)


def clear_sheet_data(sheet, column_count: int) -> None:
    amount_col_indices = {
        idx for idx, name in enumerate(EXPECTED_COLUMNS, start=1) if name in AMOUNT_COLUMNS
    }
    for row_idx in range(2, sheet.max_row + 1):
        for col_idx in range(1, column_count + 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            cell.value = None
            if col_idx in amount_col_indices:
                cell.number_format = "General"


def write_dataframe_to_sheet(sheet, df: pd.DataFrame) -> None:
    clear_sheet_data(sheet, len(EXPECTED_COLUMNS))

    for row_offset, row in enumerate(df.itertuples(index=False, name=None), start=2):
        for col_idx, (column_name, raw_value) in enumerate(
            zip(EXPECTED_COLUMNS, row), start=1
        ):
            cell = sheet.cell(row=row_offset, column=col_idx)
            cell.value = _cell_value(column_name, raw_value)
            if column_name in AMOUNT_COLUMNS and cell.value is not None:
                cell.number_format = AMOUNT_NUMBER_FORMAT


def move_pending_files(pending_files: list[Path], completed_dir: Path) -> int:
    completed_dir.mkdir(parents=True, exist_ok=True)
    moved = 0

    for src in pending_files:
        dest = completed_dir / src.name
        if dest.exists():
            dest.unlink()
        shutil.move(str(src), str(dest))
        moved += 1

    return moved


def print_summary(
    total_csv: int,
    pending_count: int,
    completed_count: int,
    row_count: int,
    moved_count: int,
    csv_dir: Path,
    xlsx_path: Path,
    *,
    dry_run: bool,
) -> None:
    print(f"읽은 CSV: {total_csv}개 (pending {pending_count}, completed {completed_count})")
    print(f"적재 행 수: {row_count}")
    if dry_run:
        print("dry-run: Excel 저장 및 파일 이동을 수행하지 않았습니다.")
        return
    if moved_count:
        print(f"이동 파일: {moved_count}개 → {csv_dir / 'completed'}")
    else:
        print("이동 파일: 0개")
    print(f"저장 완료: {xlsx_path}")


def run(
    csv_dir: Path,
    xlsx_path: Path,
    sheet_name: str,
    *,
    dry_run: bool = False,
) -> None:
    if not xlsx_path.is_file():
        raise CsvImportError(f"Excel 파일이 존재하지 않습니다: {xlsx_path}")

    pending, completed = collect_csv_files(csv_dir)
    all_files = sorted(pending + completed, key=lambda p: p.name)
    merged = read_and_merge_csv_files(all_files)

    if dry_run:
        print_summary(
            len(all_files),
            len(pending),
            len(completed),
            len(merged),
            0,
            csv_dir,
            xlsx_path,
            dry_run=True,
        )
        return

    try:
        workbook = load_workbook(xlsx_path)
    except PermissionError as exc:
        raise CsvImportError(
            f"Excel 파일을 열 수 없습니다. 파일이 다른 프로그램에서 열려 있는지 확인하세요: {xlsx_path}"
        ) from exc

    if sheet_name not in workbook.sheetnames:
        workbook.close()
        raise CsvImportError(f"시트를 찾을 수 없습니다: {sheet_name}")

    sheet = workbook[sheet_name]
    write_dataframe_to_sheet(sheet, merged)

    try:
        workbook.save(xlsx_path)
    except PermissionError as exc:
        workbook.close()
        raise CsvImportError(
            f"Excel 파일을 저장할 수 없습니다. 파일을 닫은 뒤 다시 실행하세요: {xlsx_path}"
        ) from exc
    finally:
        workbook.close()

    moved_count = move_pending_files(pending, csv_dir / "completed")
    print_summary(
        len(all_files),
        len(pending),
        len(completed),
        len(merged),
        moved_count,
        csv_dir,
        xlsx_path,
        dry_run=False,
    )


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="CSV 파일을 매출.xlsx의 매출 시트에 적재합니다."
    )
    parser.add_argument(
        "--csv-dir",
        default="resData/csv",
        help="CSV 입력 폴더 (기본값: resData/csv)",
    )
    parser.add_argument(
        "--xlsx-path",
        default="resData/매출.xlsx",
        help="대상 Excel 파일 (기본값: resData/매출.xlsx)",
    )
    parser.add_argument(
        "--sheet-name",
        default="매출",
        help="적재 대상 시트명 (기본값: 매출)",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Excel 저장 및 파일 이동 없이 검증/건수만 출력",
    )
    parser.add_argument(
        "--no-gui",
        action="store_true",
        help="GUI 없이 CLI 경로 인자만 사용",
    )
    return parser


def main(argv: list[str] | None = None) -> int:
    parser = build_parser()
    args = parser.parse_args(argv)

    if args.no_gui:
        csv_dir = resolve_path(args.csv_dir)
        xlsx_path = resolve_path(args.xlsx_path)
    else:
        csv_dir, xlsx_path = select_paths_via_gui()
        if csv_dir is None or xlsx_path is None:
            print("오류: 폴더 또는 파일 선택이 취소되었습니다.", file=sys.stderr)
            return 1

    try:
        run(
            csv_dir,
            xlsx_path,
            args.sheet_name,
            dry_run=args.dry_run,
        )
    except CsvImportError as exc:
        print(f"오류: {exc}", file=sys.stderr)
        return 1

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
